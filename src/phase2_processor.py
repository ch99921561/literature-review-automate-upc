# -*- coding: utf-8 -*-
"""
phase2_processor.py - Procesador para obtener abstracts de publicaciones
"""

import os
import re
import time
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from urllib.parse import quote

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from deep_translator import GoogleTranslator

from .config import APIType, API_CONFIGS, OUTPUTS_DIR
from .http_client import HTTPClient
from .scopus_client import ScopusAPIClient
from .ieee_client import IEEEAPIClient
from .wos_client import WOSAPIClient


class Phase2Processor:
    """Procesador para Phase 2: obtención de abstracts y traducción."""
    
    def __init__(self):
        self.http = HTTPClient()
        self.scopus_client: Optional[ScopusAPIClient] = None
        self.ieee_client: Optional[IEEEAPIClient] = None
        self.wos_client: Optional[WOSAPIClient] = None
        self.translator = GoogleTranslator(source='en', target='es')
        
        # Inicializar clientes
        self._init_clients()
    
    def _init_clients(self) -> None:
        """Inicializa los clientes de APIs."""
        print("\n[Inicializando clientes para Phase 2]")
        
        # Scopus
        self.scopus_client = ScopusAPIClient()
        if self.scopus_client.authenticate():
            print("  ✓ Scopus autenticado")
        else:
            print("  ✗ Scopus no disponible")
            self.scopus_client = None
        
        # IEEE
        self.ieee_client = IEEEAPIClient()
        if self.ieee_client.authenticate():
            print("  ✓ IEEE autenticado")
        else:
            print("  ✗ IEEE no disponible")
            self.ieee_client = None
        
        # WoS
        self.wos_client = WOSAPIClient()
        if self.wos_client.authenticate():
            print("  ✓ Web of Science autenticado")
        else:
            print("  ✗ Web of Science no disponible")
            self.wos_client = None
    
    def process_xlsx(self, input_file: str) -> str:
        """
        Procesa el archivo xlsx de phase1 y genera un nuevo archivo con abstracts.
        
        Args:
            input_file: Ruta al archivo xlsx de entrada
        
        Returns:
            Ruta al archivo de salida generado
        """
        print(f"\n{'='*80}")
        print("  PHASE 2: OBTENCIÓN DE ABSTRACTS")
        print(f"{'='*80}")
        print(f"Archivo de entrada: {input_file}")
        
        # Cargar workbook
        wb = load_workbook(input_file)
        
        # Buscar hojas de documentos
        doc_sheets = [name for name in wb.sheetnames if name.endswith("_Documentos")]
        
        if not doc_sheets:
            print("ERROR: No se encontraron hojas de documentos en el archivo")
            return ""
        
        print(f"Hojas de documentos encontradas: {', '.join(doc_sheets)}")
        
        # Crear nuevo workbook para resultados
        wb_output = Workbook()
        ws_output = wb_output.active
        ws_output.title = "Abstracts"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        # Headers - Columnas con terna de búsqueda
        headers = ["Llave", "Terna", "Article Title", "Authors", "Source Title", "Author Keywords",
                   "API_Source", "Abstract", "Abstract_ES"]
        for col, header in enumerate(headers, 1):
            cell = ws_output.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Procesar documentos
        output_row = 2
        processed_titles = set()  # Para evitar duplicados
        
        for sheet_name in doc_sheets:
            ws = wb[sheet_name]
            print(f"\nProcesando hoja: {sheet_name}")
            
            # Leer datos (saltando header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[2]:  # Verificar que hay título
                    continue
                
                llave = row[0]
                terna = row[1] if len(row) > 1 else ""  # Keywords/Terna de búsqueda
                title = row[2]
                api_source = row[3] if len(row) > 3 else ""
                
                # Evitar duplicados
                if title in processed_titles:
                    continue
                processed_titles.add(title)
                
                print(f"  [{api_source}] Buscando: {title[:55]}...")
                
                # Obtener datos de la API de origen primero
                api_data = {'abstract': None, 'authors': None, 'source_title': None, 'keywords': None}
                fallback_apis = []  # APIs para intentar si la principal falla
                
                if api_source and 'scopus' in api_source.lower():
                    api_data = self._get_scopus_data(title)
                    fallback_apis = ['wos', 'ieee']  # Scopus no da abstracts, intentar con otras
                elif api_source and 'ieee' in api_source.lower():
                    api_data = self._get_ieee_data(title)
                    fallback_apis = ['wos', 'scopus']
                elif api_source and ('wos' in api_source.lower() or 'web of science' in api_source.lower()):
                    api_data = self._get_wos_data(title)
                    fallback_apis = ['ieee', 'scopus']
                else:
                    # Si no hay API_Source definida, intentar con todas
                    api_data = self._get_scopus_data(title)
                    fallback_apis = ['wos', 'ieee']
                
                # Fallback: si no hay abstract válido, intentar con otras APIs
                abstract_val = api_data.get('abstract', '')
                if not abstract_val or 'no disponible' in str(abstract_val).lower() or 'No encontrado' in str(abstract_val):
                    for fallback in fallback_apis:
                        if fallback == 'wos':
                            fallback_data = self._get_wos_data(title)
                        elif fallback == 'ieee':
                            fallback_data = self._get_ieee_data(title)
                        else:
                            fallback_data = self._get_scopus_data(title)
                        
                        fallback_abstract = fallback_data.get('abstract', '')
                        if fallback_abstract and 'no disponible' not in str(fallback_abstract).lower():
                            # Usar datos del fallback pero mantener metadatos originales si existen
                            api_data['abstract'] = fallback_abstract
                            if not api_data.get('authors'):
                                api_data['authors'] = fallback_data.get('authors')
                            if not api_data.get('source_title'):
                                api_data['source_title'] = fallback_data.get('source_title')
                            if not api_data.get('keywords'):
                                api_data['keywords'] = fallback_data.get('keywords')
                            break
                
                time.sleep(0.3)
                
                # Extraer datos
                abstract = self._normalize_text(api_data.get('abstract', ''))
                authors = api_data.get('authors') or ""
                source_title = api_data.get('source_title') or ""
                keywords = api_data.get('keywords') or ""
                
                # Traducir abstract
                abstract_es = self._translate(abstract) if abstract else ""
                
                # Escribir fila
                # Col 1: Llave
                ws_output.cell(row=output_row, column=1, value=llave).border = border
                
                # Col 2: Terna
                cell_terna = ws_output.cell(row=output_row, column=2, value=terna)
                cell_terna.border = border
                cell_terna.alignment = wrap_alignment
                
                # Col 3: Article Title
                cell_title = ws_output.cell(row=output_row, column=3, value=title)
                cell_title.border = border
                cell_title.alignment = wrap_alignment
                
                # Col 4: Authors
                cell_authors = ws_output.cell(row=output_row, column=4, value=authors or "No disponible")
                cell_authors.border = border
                cell_authors.alignment = wrap_alignment
                
                # Col 5: Source Title
                cell_source = ws_output.cell(row=output_row, column=5, value=source_title or "No disponible")
                cell_source.border = border
                cell_source.alignment = wrap_alignment
                
                # Col 6: Author Keywords
                cell_keywords = ws_output.cell(row=output_row, column=6, value=keywords or "No disponible")
                cell_keywords.border = border
                cell_keywords.alignment = wrap_alignment
                
                # Col 7: API_Source
                ws_output.cell(row=output_row, column=7, value=api_source).border = border
                
                # Col 8: Abstract
                cell_abstract = ws_output.cell(row=output_row, column=8, value=abstract or "No encontrado")
                cell_abstract.border = border
                cell_abstract.alignment = wrap_alignment
                
                # Col 9: Abstract_ES
                cell_es = ws_output.cell(row=output_row, column=9, value=abstract_es or "No disponible")
                cell_es.border = border
                cell_es.alignment = wrap_alignment
                
                output_row += 1
                
                # Progreso
                if output_row % 10 == 0:
                    print(f"    Procesados: {output_row - 2} documentos")
        
        # Ajustar anchos de columna
        ws_output.column_dimensions['A'].width = 8    # Llave
        ws_output.column_dimensions['B'].width = 50   # Terna
        ws_output.column_dimensions['C'].width = 60   # Article Title
        ws_output.column_dimensions['D'].width = 40   # Authors
        ws_output.column_dimensions['E'].width = 40   # Source Title
        ws_output.column_dimensions['F'].width = 35   # Author Keywords
        ws_output.column_dimensions['G'].width = 12   # API_Source
        ws_output.column_dimensions['H'].width = 80   # Abstract
        ws_output.column_dimensions['I'].width = 80   # Abstract_ES
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(OUTPUTS_DIR, f"abstracts_{timestamp}.xlsx")
        wb_output.save(output_file)
        
        print(f"\n{'='*80}")
        print(f"  PHASE 2 COMPLETADA")
        print(f"{'='*80}")
        print(f"Documentos procesados: {output_row - 2}")
        print(f"Archivo generado: {output_file}")
        
        return output_file
    
    def _clean_title_for_search(self, title: str) -> str:
        """
        Limpia el título para búsquedas API removiendo caracteres problemáticos.
        
        Caracteres que causan problemas:
        - ?, !, :, ;, . - Pueden ser interpretados incorrectamente
        - (, ), [, ] - Caracteres de agrupación
        - ", ', ', ' - Comillas y apóstrofes (causan "China's" -> "Chinas")
        - –, —, - - Guiones especiales (em/en dash) y guión normal
        - \xa0 - Espacio no rompible
        - © - Símbolo de copyright
        """
        # 1. Reemplazar caracteres especiales por espacios (guión al final de la clase)
        title_clean = re.sub(r'[?!:;,\.\(\)\[\]"\'©–—''"\xa0-]', ' ', title)
        
        # 2. Normalizar posesivos: "Chinas" -> "China" (ocurre cuando se remueve apóstrofe)
        title_clean = re.sub(r'\b([A-Z][a-z]+a)s\b', r'\1', title_clean)
        
        # 3. Normalizar espacios múltiples
        title_clean = re.sub(r'\s+', ' ', title_clean).strip()
        
        return title_clean
    
    def _get_scopus_data(self, title: str) -> Dict:
        """Obtiene datos de Scopus: abstract, authors, source, keywords.
        
        NOTA: La API de Scopus con API key estándar NO incluye abstracts.
        Para obtener abstracts se requiere view=COMPLETE que necesita suscripción premium.
        Se obtienen metadatos básicos disponibles con view=STANDARD.
        """
        result = {'abstract': None, 'authors': None, 'source_title': None, 'keywords': None}
        
        if not self.scopus_client:
            return result
        
        try:
            title_clean = self._clean_title_for_search(title)
            query = f'TITLE("{title_clean}")'
            config = API_CONFIGS[APIType.SCOPUS]
            
            # Búsqueda con view=STANDARD (único disponible con API key estándar)
            url = f"{config.base_url}?query={quote(query)}&count=1&view=STANDARD"
            
            headers = {"X-ELS-APIKey": self.scopus_client.api_key}
            response = self.http.get(url, headers=headers, verbose=False)
            
            if "error" in response:
                return result
            
            entries = response.get("search-results", {}).get("entry", [])
            if not entries or len(entries) == 0:
                return result
            
            entry = entries[0]
            
            # Metadatos básicos disponibles con view=STANDARD
            result['authors'] = entry.get("dc:creator", "")
            result['source_title'] = entry.get("prism:publicationName", "")
            
            # Abstract NO disponible con API key estándar - dejamos None para fallback a otras APIs
            # result['abstract'] permanece None
            
            # Keywords tampoco disponibles con view=STANDARD
            # result['keywords'] permanece None
                    
        except Exception as e:
            pass
        
        return result
    
    def _get_ieee_data(self, title: str) -> Dict:
        """Obtiene datos completos de IEEE: abstract, authors, source, keywords."""
        result = {'abstract': None, 'authors': None, 'source_title': None, 'keywords': None}
        
        if not self.ieee_client:
            return result
        
        try:
            title_clean = self._clean_title_for_search(title)
            config = API_CONFIGS[APIType.IEEE]
            
            # Búsqueda por título
            url = f"{config.base_url}?apikey={self.ieee_client.api_key}&article_title={quote(title_clean)}&max_records=3"
            response = self.http.get(url, verbose=False, mask_key=self.ieee_client.api_key)
            
            article = None
            if "error" not in response:
                articles = response.get("articles", [])
                if articles:
                    article = articles[0]
            
            # Intento 2: Búsqueda por palabras clave si no encontró
            if not article:
                stopwords = {'and', 'or', 'not', 'the', 'of', 'in', 'to', 'a', 'an', 
                            'for', 'on', 'with', 'by', 'from', 'as', 'at', 'under', 'near'}
                words = [w for w in title_clean.split() if w.lower() not in stopwords and len(w) > 2]
                
                if len(words) >= 3:
                    search_terms = ' '.join(words[:5])
                    url = f"{config.base_url}?apikey={self.ieee_client.api_key}&querytext={quote(search_terms)}&max_records=5"
                    response = self.http.get(url, verbose=False, mask_key=self.ieee_client.api_key)
                    
                    if "error" not in response:
                        articles = response.get("articles", [])
                        for art in articles:
                            article_title = art.get("title", "").lower()
                            matches = sum(1 for w in words[:4] if w.lower() in article_title)
                            if matches >= 3:
                                article = art
                                break
            
            if article:
                # Abstract
                result['abstract'] = article.get("abstract", "")
                
                # Authors
                authors_data = article.get("authors", {}).get("authors", [])
                if isinstance(authors_data, list):
                    author_names = [a.get("full_name", "") for a in authors_data if isinstance(a, dict)]
                    result['authors'] = "; ".join(filter(None, author_names))
                
                # Source Title
                result['source_title'] = article.get("publication_title", "") or article.get("publisher", "")
                
                # Keywords - index_terms
                index_terms = article.get("index_terms", {})
                keywords_list = []
                
                # IEEE terms
                ieee_terms = index_terms.get("ieee_terms", {}).get("terms", [])
                if isinstance(ieee_terms, list):
                    keywords_list.extend(ieee_terms)
                
                # Author terms
                author_terms = index_terms.get("author_terms", {}).get("terms", [])
                if isinstance(author_terms, list):
                    keywords_list.extend(author_terms)
                
                result['keywords'] = "; ".join(keywords_list) if keywords_list else ""
                
        except Exception as e:
            pass
        
        return result
    
    def _get_wos_data(self, title: str) -> Dict:
        """Obtiene datos completos de WoS: abstract, authors, source, keywords."""
        result = {'abstract': None, 'authors': None, 'source_title': None, 'keywords': None}
        
        if not self.wos_client:
            return result
        
        try:
            config = API_CONFIGS[APIType.WOS]
            title_clean = self._clean_title_for_search(title)
            
            stopwords = {'and', 'or', 'not', 'the', 'of', 'in', 'to', 'a', 'an', 
                        'for', 'on', 'with', 'by', 'from', 'as', 'at', 'under', 'near'}
            words = [w for w in title_clean.split() if w.lower() not in stopwords]
            title_search = ' '.join(words[:8])
            
            query = f'TS=({title_search})'
            url = f"{config.base_url}?databaseId=WOS&usrQuery={quote(query)}&count=1&firstRecord=1"
            
            headers = {"X-ApiKey": self.wos_client.api_key}
            response = self.http.get(url, headers=headers, verbose=False)
            
            if "error" in response:
                return result
            
            records = response.get("Data", {}).get("Records", {}).get("records", {}).get("REC", [])
            if records:
                if isinstance(records, dict):
                    records = [records]
                if len(records) > 0:
                    rec = records[0]
                    static_data = rec.get("static_data", {})
                    
                    # Abstract
                    fullrecord_metadata = static_data.get("fullrecord_metadata", {})
                    abstracts = fullrecord_metadata.get("abstracts", {}).get("abstract", {})
                    
                    if isinstance(abstracts, dict):
                        text = abstracts.get("abstract_text", {})
                        if isinstance(text, dict):
                            p = text.get("p", "")
                            result['abstract'] = " ".join(str(item) for item in p if item) if isinstance(p, list) else p
                        elif isinstance(text, str):
                            result['abstract'] = text
                    elif isinstance(abstracts, list) and len(abstracts) > 0:
                        text = abstracts[0].get("abstract_text", {})
                        if isinstance(text, dict):
                            p = text.get("p", "")
                            result['abstract'] = " ".join(str(item) for item in p if item) if isinstance(p, list) else p
                        elif isinstance(text, str):
                            result['abstract'] = text
                    
                    # Authors
                    summary = static_data.get("summary", {})
                    names = summary.get("names", {}).get("name", [])
                    if isinstance(names, dict):
                        names = [names]
                    author_names = []
                    for name in names:
                        if name.get("role") == "author":
                            full_name = name.get("full_name", "") or name.get("display_name", "")
                            if full_name:
                                author_names.append(full_name)
                    result['authors'] = "; ".join(author_names)
                    
                    # Source Title
                    titles = summary.get("titles", {}).get("title", [])
                    if isinstance(titles, dict):
                        titles = [titles]
                    for t in titles:
                        if t.get("type") == "source":
                            result['source_title'] = t.get("content", "")
                            break
                    
                    # Keywords
                    keywords_data = fullrecord_metadata.get("keywords", {}).get("keyword", [])
                    if isinstance(keywords_data, list):
                        result['keywords'] = "; ".join(keywords_data)
                    elif isinstance(keywords_data, str):
                        result['keywords'] = keywords_data
                        
        except Exception as e:
            pass
        
        return result
    
    def _get_scopus_abstract(self, title: str) -> Optional[str]:
        """Obtiene el abstract de Scopus por título."""
        if not self.scopus_client:
            return None
        
        try:
            # Scopus es más tolerante, pero limpiamos caracteres problemáticos
            title_clean = self._clean_title_for_search(title)
            
            # Buscar por título - Scopus acepta búsqueda con comillas
            # Agregamos view=COMPLETE para obtener dc:description (abstract)
            query = f'TITLE("{title_clean}")'
            config = API_CONFIGS[APIType.SCOPUS]
            url = f"{config.base_url}?query={quote(query)}&count=1&view=COMPLETE"
            
            headers = {"X-ELS-APIKey": self.scopus_client.api_key}
            response = self.http.get(url, headers=headers, verbose=False)
            
            if "error" in response:
                return None
            
            # Extraer abstract
            entries = response.get("search-results", {}).get("entry", [])
            if entries and len(entries) > 0:
                abstract = entries[0].get("dc:description", "")
                if abstract:
                    return abstract
                
                # Si no hay abstract en dc:description, intentar con prism:description
                abstract = entries[0].get("prism:description", "")
                return abstract if abstract else None
        except Exception as e:
            pass
        
        return None
    
    def _get_ieee_abstract(self, title: str) -> Optional[str]:
        """Obtiene el abstract de IEEE por título."""
        if not self.ieee_client:
            return None
        
        try:
            # Limpiar título para IEEE
            title_clean = self._clean_title_for_search(title)
            
            config = API_CONFIGS[APIType.IEEE]
            
            # Intento 1: Búsqueda por título completo
            url = f"{config.base_url}?apikey={self.ieee_client.api_key}&article_title={quote(title_clean)}&max_records=3"
            response = self.http.get(url, verbose=False, mask_key=self.ieee_client.api_key)
            
            if "error" not in response:
                articles = response.get("articles", [])
                if articles and len(articles) > 0:
                    abstract = articles[0].get("abstract", "")
                    if abstract:
                        return abstract
            
            # Intento 2: Búsqueda por palabras clave del título
            # Útil para títulos con caracteres especiales como "SoK:"
            stopwords = {'and', 'or', 'not', 'the', 'of', 'in', 'to', 'a', 'an', 
                        'for', 'on', 'with', 'by', 'from', 'as', 'at', 'under', 'near'}
            words = [w for w in title_clean.split() if w.lower() not in stopwords and len(w) > 2]
            
            if len(words) >= 3:
                # Usar las primeras palabras distintivas
                search_terms = ' '.join(words[:5])
                url = f"{config.base_url}?apikey={self.ieee_client.api_key}&querytext={quote(search_terms)}&max_records=5"
                response = self.http.get(url, verbose=False, mask_key=self.ieee_client.api_key)
                
                if "error" not in response:
                    articles = response.get("articles", [])
                    # Buscar el artículo que coincida mejor con el título
                    for article in articles:
                        article_title = article.get("title", "").lower()
                        # Verificar si las palabras clave principales están en el título
                        matches = sum(1 for w in words[:4] if w.lower() in article_title)
                        if matches >= 3:
                            abstract = article.get("abstract", "")
                            if abstract:
                                return abstract
        except Exception as e:
            pass
        
        return None
    
    def _get_wos_abstract(self, title: str) -> Optional[str]:
        """Obtiene el abstract de Web of Science por título."""
        if not self.wos_client:
            return None
        
        try:
            # WoS Starter API - buscar por título
            # Usamos TS (Topic Search) en lugar de TI (Title) porque TI falla
            # con caracteres especiales como ?, :, !, etc.
            config = API_CONFIGS[APIType.WOS]
            
            # 1. Limpiar título usando método centralizado
            title_clean = self._clean_title_for_search(title)
            
            # 2. Remover stopwords que WoS interpreta como operadores booleanos
            stopwords = {'and', 'or', 'not', 'the', 'of', 'in', 'to', 'a', 'an', 
                        'for', 'on', 'with', 'by', 'from', 'as', 'at', 'under', 'near'}
            words = [w for w in title_clean.split() if w.lower() not in stopwords]
            
            # 3. Tomar las primeras palabras significativas (máx 8)
            title_search = ' '.join(words[:8])
            
            query = f'TS=({title_search})'
            url = f"{config.base_url}?databaseId=WOS&usrQuery={quote(query)}&count=1&firstRecord=1"
            
            headers = {"X-ApiKey": self.wos_client.api_key}
            response = self.http.get(url, headers=headers, verbose=False)
            
            if "error" in response:
                return None
            
            # Extraer abstract - WoS tiene estructura diferente
            records = response.get("Data", {}).get("Records", {}).get("records", {}).get("REC", [])
            if records:
                if isinstance(records, dict):
                    records = [records]
                if len(records) > 0:
                    # El abstract puede estar en diferentes ubicaciones según la respuesta
                    static_data = records[0].get("static_data", {})
                    fullrecord_metadata = static_data.get("fullrecord_metadata", {})
                    abstracts = fullrecord_metadata.get("abstracts", {}).get("abstract", {})
                    
                    if isinstance(abstracts, dict):
                        text = abstracts.get("abstract_text", {})
                        if isinstance(text, dict):
                            # p puede ser lista de párrafos o string
                            p = text.get("p", "")
                            if isinstance(p, list):
                                return " ".join(str(item) for item in p if item)
                            return p
                        return text if isinstance(text, str) else None
                    elif isinstance(abstracts, list) and len(abstracts) > 0:
                        text = abstracts[0].get("abstract_text", {})
                        if isinstance(text, dict):
                            p = text.get("p", "")
                            if isinstance(p, list):
                                return " ".join(str(item) for item in p if item)
                            return p
                        return text if isinstance(text, str) else None
        except Exception as e:
            pass
        
        return None
    
    def _normalize_text(self, text) -> str:
        """Normaliza el texto (convierte listas a string)."""
        if text is None:
            return ""
        if isinstance(text, list):
            # Unir elementos de lista
            return " ".join(str(item) for item in text if item)
        if isinstance(text, dict):
            # Intentar extraer texto de dict
            return str(text.get("p", "") or text.get("text", "") or "")
        return str(text)
    
    def _translate(self, text) -> str:
        """Traduce texto de inglés a español."""
        # Normalizar texto primero
        text = self._normalize_text(text)
        
        if not text or len(text.strip()) == 0:
            return ""
        
        try:
            # deep-translator tiene límite de caracteres, dividir si es necesario
            max_chars = 4500
            if len(text) <= max_chars:
                return self.translator.translate(text)
            else:
                # Dividir en partes
                parts = []
                for i in range(0, len(text), max_chars):
                    part = text[i:i+max_chars]
                    translated = self.translator.translate(part)
                    parts.append(translated)
                    time.sleep(0.2)
                return " ".join(parts)
        except Exception as e:
            return f"[Error de traducción: {str(e)}]"


def run_phase2(input_file: str) -> int:
    """
    Ejecuta Phase 2.
    
    Args:
        input_file: Ruta al archivo xlsx de entrada
    
    Returns:
        Código de retorno (0 = éxito)
    """
    processor = Phase2Processor()
    output_file = processor.process_xlsx(input_file)
    
    if output_file:
        return 0
    return 1
