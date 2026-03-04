# -*- coding: utf-8 -*-
"""Analizar archivo anterior que SÍ tenía abstracts."""
from openpyxl import load_workbook

wb = load_workbook('outputs/abstracts_20260221_172700.xlsx')
ws = wb.active

print(f'Headers: {[cell.value for cell in ws[1]]}')
print(f'Total rows: {ws.max_row}')

# Headers: Llave, Titulo, API_Source, Abstract_Scopus, Abstract_IEEE, Abstract_WoS, Abstract_ES
# Indices:  0      1       2           3                4              5            6

# Ver primeras filas completas
print('\n=== PRIMERAS 3 FILAS ===')
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
    print(f'Row {i+2}:')
    print(f'  Llave: {row[0]}')
    print(f'  Titulo: {str(row[1])[:50]}...')
    print(f'  API_Source: {row[2]}')
    print(f'  Abstract_Scopus: {str(row[3])[:80] if row[3] else "VACIO"}...')
    print(f'  Abstract_IEEE: {str(row[4])[:80] if row[4] else "VACIO"}...')
    print(f'  Abstract_WoS: {str(row[5])[:80] if row[5] else "VACIO"}...')
    print(f'  Abstract_ES: {str(row[6])[:80] if row[6] else "VACIO"}...')
    print()

# Contar documentos con algún abstract
print('=== ESTADÍSTICAS POR API_SOURCE ===')
for api_src in ['scopus', 'wos', 'ieee']:
    total = 0
    with_abs = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        source = str(row[2] or '').lower()
        if api_src in source:
            total += 1
            # Verificar si tiene algún abstract (cualquier columna)
            has_abs = False
            for col in [3, 4, 5, 6]:  # Abstract_Scopus, IEEE, WoS, ES
                val = row[col] if len(row) > col else None
                if val and 'No encontrado' not in str(val) and len(str(val)) > 20:
                    has_abs = True
                    break
            if has_abs:
                with_abs += 1
    if total > 0:
        print(f'{api_src.upper()}: {with_abs}/{total} con algún abstract ({100*with_abs/total:.1f}%)')
